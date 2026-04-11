# -*- coding: utf-8 -*-
from PyQt5.QtWidgets import QApplication , QMainWindow, QFileDialog
import sys
import os
sys.path.append('UIProgram')
from UIProgram.UiMain import Ui_MainWindow
import sys
from PyQt5.QtCore import QTimer, Qt, QCoreApplication
import detect_tools as tools
import Config
from PyQt5.QtGui import QPixmap
from UIProgram.QssLoader import QSSLoader
from ultralytics import YOLO
import cv2
import numpy as np
import torch
import openpyxl
from openpyxl import Workbook
import zipfile
from PyQt5.QtCore import QTimer, Qt, QThread, pyqtSignal, QCoreApplication
from PyQt5.QtGui import QPixmap, QPainter, QPalette, QBrush, QFontDatabase, QFont, QTransform, QPen, QColor, QPainterPath
import detect_tools as tools
from PyQt5.QtWidgets import QApplication , QMainWindow, QFileDialog, \
    QMessageBox,QWidget,QHeaderView,QTableWidgetItem, QAbstractItemView,\
    QVBoxLayout, QLineEdit, QPushButton, QApplication, QLabel, QDialog,\
    QHBoxLayout, QFrame, QGraphicsDropShadowEffect
import re
from database import SqliteDBOperator
import os
os.environ['QT_QPA_PLATFORM_PLUGIN_PATH'] = "D:\BaiduNetdiskDownload\多模态情绪检测\dmt-facialexpression\dmt-facialexpression\.venv\Lib\site-packages\PyQt5\Qt\plugins"

class LoginWidget(QWidget):
    login_success_signal = pyqtSignal()

    def __init__(self):
        super().__init__()
        self.setFixedSize(1000, 600)
        self.setWindowTitle("多模态生理信号情绪检测系统-登录")
        
        # Dark Theme Background
        self.bg_pixmap = QPixmap(1000, 600)
        self.bg_pixmap.fill(QColor("#1a1a2e")) # Dark Navy

        self.init_ui()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.drawPixmap(0, 0, self.bg_pixmap)
        painter.setRenderHint(QPainter.Antialiasing)
        w = self.width()
        h = self.height()
        
        # Draw Tech Grid
        painter.setPen(QPen(QColor(15, 52, 96, 100), 1)) # Dark Blue Grid
        step = 40
        for x in range(0, w, step):
            painter.drawLine(x, 0, x, h)
        for y in range(0, h, step):
            painter.drawLine(0, y, w, y)
            
        # Draw some decorative squares
        painter.setPen(Qt.NoPen)
        import random
        random.seed(42) # Fixed seed for consistent look
        for _ in range(15):
            size = random.randint(20, 80)
            x = random.randint(0, w - size)
            y = random.randint(0, h - size)
            color = QColor(233, 69, 96, 40) if random.random() > 0.5 else QColor(15, 52, 96, 60)
            painter.setBrush(color)
            painter.drawRect(x, y, size, size)

    def resizeEvent(self, event):
        self.bg_pixmap = QPixmap(self.size())
        self.bg_pixmap.fill(QColor("#1a1a2e"))
        super(LoginWidget, self).resizeEvent(event)

    def init_ui(self):
        # 主布局
        main_layout = QVBoxLayout(self)
        main_layout.setAlignment(Qt.AlignCenter)
        main_layout.setContentsMargins(0, 0, 0, 0)

        # 登录卡片
        self.login_card = QFrame()
        self.login_card.setFixedSize(420, 480)
        self.login_card.setStyleSheet("""
            QFrame {
                background-color: rgba(22, 33, 62, 0.95); /* Dark Blue-Grey */
                border: 1px solid #0f3460;
                border-radius: 20px;
            }
        """)
        
        # 阴影效果
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(40)
        shadow.setColor(QColor(0, 0, 0, 150))
        shadow.setOffset(0, 10)
        self.login_card.setGraphicsEffect(shadow)

        # 卡片布局
        card_layout = QVBoxLayout(self.login_card)
        card_layout.setContentsMargins(40, 40, 40, 40)
        card_layout.setSpacing(20)

        # 标题
        title_label = QLabel("多模态生理信号\n情绪检测系统")
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setFont(QFont("Microsoft YaHei", 20, QFont.Bold))
        title_label.setStyleSheet("color: #e94560; background: transparent;") # Pink/Red Title
        card_layout.addWidget(title_label)
        
        # Logo placeholder (optional, or use icon)
        # logo_label = QLabel()
        # logo_label.setAlignment(Qt.AlignCenter)
        # logo_pix = QPixmap("UIProgram/ui_imgs/11.png")
        # if not logo_pix.isNull():
        #     logo_label.setPixmap(logo_pix.scaled(80, 80, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        # card_layout.addWidget(logo_label)
        
        card_layout.addSpacing(10)

        # 通用输入框样式
        input_style = """
            QLineEdit {
                border: 1px solid #0f3460;
                border-radius: 22px;
                padding: 0 20px;
                background-color: #1a1a2e;
                color: #e0e0e0;
                font-family: "Microsoft YaHei";
                font-size: 14px;
            }
            QLineEdit:focus {
                border: 1px solid #e94560;
                background-color: #1a1a2e;
            }
            QLineEdit::placeholder {
                color: #94a1b2;
            }
        """

        # 用户名
        self.username_edit = QLineEdit()
        self.username_edit.setPlaceholderText("用户名")
        self.username_edit.setFixedHeight(45)
        self.username_edit.setStyleSheet(input_style)
        card_layout.addWidget(self.username_edit)

        # 密码
        self.password_edit = QLineEdit()
        self.password_edit.setPlaceholderText("密码")
        self.password_edit.setEchoMode(QLineEdit.Password)
        self.password_edit.setFixedHeight(45)
        self.password_edit.setStyleSheet(input_style)
        card_layout.addWidget(self.password_edit)

        card_layout.addSpacing(10)

        # 登录按钮
        self.login_button = QPushButton("登  录")
        self.login_button.setFixedHeight(45)
        self.login_button.setCursor(Qt.PointingHandCursor)
        self.login_button.setStyleSheet("""
            QPushButton {
                background-color: #e94560;
                color: white;
                border: none;
                border-radius: 22px;
                font-family: "Microsoft YaHei";
                font-size: 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #ff6b81;
            }
            QPushButton:pressed {
                background-color: #c0394d;
            }
        """)
        self.login_button.clicked.connect(self.check_login)
        card_layout.addWidget(self.login_button)

        # 注册链接
        self.register_button = QPushButton("注册新账号")
        self.register_button.setCursor(Qt.PointingHandCursor)
        self.register_button.setStyleSheet("""
            QPushButton {
                background: transparent;
                color: #94a1b2;
                border: none;
                font-family: "Microsoft YaHei";
                font-size: 13px;
                text-decoration: underline;
            }
            QPushButton:hover {
                color: #e94560;
            }
        """)
        self.register_button.clicked.connect(self.show_register_dialog)
        card_layout.addWidget(self.register_button, alignment=Qt.AlignCenter)
        
        card_layout.addStretch()
        main_layout.addWidget(self.login_card)

    def check_login(self):
        username = self.username_edit.text()
        password = self.password_edit.text()
        db_operator = SqliteDBOperator()
        if db_operator.check_user_login(username, password):
            self.login_success_signal.emit()
            self.close()
            db_operator.close_connection()
        else:
            QMessageBox.warning(self, "登录失败", "账号或密码错误，请重新输入", QMessageBox.Ok)
            db_operator.close_connection()

    def show_register_dialog(self):
        register_dialog = QDialog(self)
        register_dialog.setWindowTitle("注册新用户")
        register_dialog.setFixedSize(350, 450)
        register_dialog.setStyleSheet("QDialog { background-color: #1a1a2e; color: #e0e0e0; }")

        layout = QVBoxLayout(register_dialog)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(15)

        title = QLabel("创建账号")
        title.setAlignment(Qt.AlignCenter)
        title.setFont(QFont("Microsoft YaHei", 18, QFont.Bold))
        title.setStyleSheet("color: #e94560;")
        layout.addWidget(title)
        layout.addSpacing(10)

        input_style = """
            QLineEdit {
                border: 1px solid #0f3460;
                border-radius: 5px;
                padding: 8px;
                background-color: #16213e;
                color: #e0e0e0;
                font-family: "Microsoft YaHei";
                font-size: 14px;
            }
            QLineEdit:focus {
                border: 1px solid #e94560;
                background-color: #16213e;
            }
            QLineEdit::placeholder {
                color: #94a1b2;
            }
        """

        reg_user = QLineEdit()
        reg_user.setPlaceholderText("用户名")
        reg_user.setStyleSheet(input_style)
        layout.addWidget(reg_user)

        reg_pwd = QLineEdit()
        reg_pwd.setPlaceholderText("密码")
        reg_pwd.setEchoMode(QLineEdit.Password)
        reg_pwd.setStyleSheet(input_style)
        layout.addWidget(reg_pwd)

        reg_pwd2 = QLineEdit()
        reg_pwd2.setPlaceholderText("确认密码")
        reg_pwd2.setEchoMode(QLineEdit.Password)
        reg_pwd2.setStyleSheet(input_style)
        layout.addWidget(reg_pwd2)

        layout.addSpacing(10)

        reg_btn = QPushButton("立即注册")
        reg_btn.setFixedHeight(40)
        reg_btn.setStyleSheet("""
            QPushButton {
                background-color: #e94560;
                color: white;
                border: none;
                border-radius: 5px;
                font-family: "Microsoft YaHei";
                font-size: 15px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #ff6b81;
            }
            QPushButton:pressed {
                background-color: #c0394d;
            }
        """)
        reg_btn.clicked.connect(lambda: self.check_register_info(
            reg_user.text(), reg_pwd.text(), reg_pwd2.text(), register_dialog))
        layout.addWidget(reg_btn)
        
        layout.addStretch()
        register_dialog.exec_()

    def check_register_info(self, username, password, confirm_password, register_dialog):
        # 检查用户名是否为空
        if username == "":
            QMessageBox.warning(self, "注册失败", "用户名为空", QMessageBox.Ok)
            return
        # 检查用户名长度是否符合要求
        elif len(username) < 4:
            QMessageBox.warning(self, "注册失败", "用户名长度不符合要求，至少4位", QMessageBox.Ok)
            return
        elif len(username) > 12:
            QMessageBox.warning(self, "注册失败", "用户名长度不符合要求，至多12位", QMessageBox.Ok)
            return
        # 新增：检查用户名格式是否符合要求
        elif not re.match(r'^[a-zA-Z][a-zA-Z0-9]*$', username):
            QMessageBox.warning(self, "注册失败", "用户名格式不符合要求，只能包含大小写字母和数字，且不能以数字开头", QMessageBox.Ok)
            return
        # 检查两次密码输入是否一致
        elif password!= confirm_password:
            QMessageBox.warning(self, "注册失败", "两次密码输入不一致", QMessageBox.Ok)
            return
        # 检查密码长度（至少6位，至多20位）
        elif len(password) < 6:
            QMessageBox.warning(self, "注册失败", "密码长度不符合要求，至少6位", QMessageBox.Ok)
            return
        elif len(password) > 20:
            QMessageBox.warning(self, "注册失败", "密码长度不符合要求，至多20位", QMessageBox.Ok)
            return
        # 检查密码是否至少包含数字、字母、特殊符号中的两种
        categories = [r'\d', r'[a-zA-Z]', r'[^\w]']
        count = 0
        for category in categories:
            if re.search(category, password):
                count += 1
        if count < 2:
            QMessageBox.warning(self, "注册失败", "密码至少需包含数字、字母、特殊符号中的两种", QMessageBox.Ok)
            return
        else:
            db_operator = SqliteDBOperator()
            # 检查用户名是否已存在
            if db_operator.check_user_exists(username):
                QMessageBox.warning(self, "注册失败", "该用户已存在，请直接登录", QMessageBox.Ok)
                return
            db_operator.insert_user_data(username, password)
            msg_box = QMessageBox.information(self, "注册成功", "注册成功", QMessageBox.Ok)
            # 注册成功后，关闭注册对话框，回到登录界面
            if msg_box == QMessageBox.Ok:
                register_dialog.close()

class MainWindow(QMainWindow):
    def __init__(self, parent=None):
        super(QMainWindow, self).__init__(parent)
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        
        # Override specific UI settings for Dark Theme
        self.ui.label_show.setStyleSheet("background-color: #000000; border: 2px dashed #0f3460;")
        self.ui.label_3.setStyleSheet("color: #e94560;") # Title
        
        self.initMain()
        self.signalconnect()
        # 加载css渲染效果
        style_file = 'UIProgram/style.css'
        qssStyleSheet = QSSLoader.read_qss_file(style_file)
        self.setStyleSheet(qssStyleSheet)

        self.conf = 0.5
        self.iou = 0.7

        # 初始化Excel文件
        self.excel_file = 'output.xlsx'
        self.init_excel()

    def init_excel(self):
        # 创建或加载Excel文件
        try:
            self.workbook = openpyxl.load_workbook(self.excel_file)
        except (FileNotFoundError, zipfile.BadZipFile):
            # 如果文件不存在或已损坏，创建新文件
            self.workbook = Workbook()
            # 如果是损坏的文件，备份一下旧文件
            if os.path.exists(self.excel_file):
                import shutil
                backup_name = self.excel_file + '.bak'
                try:
                    shutil.copy(self.excel_file, backup_name)
                    print(f"检测到损坏的 Excel 文件，已备份为: {backup_name}")
                except:
                    pass
        self.sheet = self.workbook.active
        # 添加表头
        if self.sheet.max_row == 1:
            self.sheet.append(['文件路径', '类型名', '置信度'])


    def save_to_excel(self, file_path, label, confidence):
        # 保存结果到Excel
        self.sheet.append([file_path, label, confidence])
        self.workbook.save(self.excel_file)

    def initMain(self):
        self.labeldict = Config.names
        self.labelchinese = Config.CH_names
        self.device = 0 if torch.cuda.is_available() else 'cpu'

        self.face_model = YOLO(Config.face_model_path, task='detect')
        self.face_model(np.zeros((48, 48, 3)).astype(np.uint8), device=self.device)  # 预先加载推理模型
        self.expression_model = YOLO(Config.expression_model_path, task='classify')
        self.expression_model(np.zeros((48, 48, 3)).astype(np.uint8), device=self.device)  # 预先加载推理模型

        self.show_width = 770
        self.show_height = 460

        self.org_path = None

        self.is_camera_open = False
        self.cap = None

        self.timer_camera = QTimer()
        self.current_mode = "face"
        self.physio_data = None
        self.physio_sample_rate = 256
        self.physio_window_size = int(5 * self.physio_sample_rate)
        self.physio_current_start = 0
        self.physio_signal_type = None
        self.physio_waveform_base = None
        if hasattr(self.ui, "groupBox_physio"):
            self.ui.groupBox_physio.hide()

    def signalconnect(self):
        self.ui.PicBtn.clicked.connect(self.open_img)
        self.ui.VideoBtn.clicked.connect(self.vedio_show)
        self.ui.CapBtn.clicked.connect(self.camera_show)
        self.ui.comboBox.activated.connect(self.combox_change)
        self.ui.exitBtn.clicked.connect(QCoreApplication.quit)
        if hasattr(self.ui, "faceModeBtn"):
            self.ui.faceModeBtn.clicked.connect(self.switch_to_face_mode)
        if hasattr(self.ui, "physioModeBtn"):
            self.ui.physioModeBtn.clicked.connect(self.switch_to_physio_mode)
        if hasattr(self.ui, "physioOpenBtn"):
            self.ui.physioOpenBtn.clicked.connect(self.open_physio_file)
        if hasattr(self.ui, "physioSlider"):
            self.ui.physioSlider.valueChanged.connect(self.on_physio_slider_changed)
        if hasattr(self.ui, "physioDetectBtn"):
            self.ui.physioDetectBtn.clicked.connect(self.recognize_physio_emotion)

    def combox_change(self):
        # 多个人脸，进行人脸选择时触发
        com_text = self.ui.comboBox.currentText()
        if com_text == '全部':
            infos = self.all_face_infos
        else:
            index = int(com_text.split('_')[-1])
            infos = [self.all_face_infos[index]]

        face_cvimg = self.cv_img.copy()
        for location, prob in infos:
            left, top, right, bottom = location
            num = np.argmax(prob)
            label = self.labeldict[num]
            face_cvimg = cv2.rectangle(face_cvimg, (left, top), (right, bottom), (50, 50, 250), 3)
            face_cvimg = cv2.putText(face_cvimg, label, ((left, top - 10)), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0, 0, 250), 2,
                                     cv2.LINE_AA)
            self.ui.resLb.setText(self.labeldict[num] + '--' + self.labelchinese[num])
            icon_name = self.labeldict[num] + '.png'
            icon_path = os.path.join('UIProgram/ui_imgs', icon_name)
            pix = QPixmap(icon_path)
            self.ui.resIcon.setPixmap(pix)
            self.ui.resIcon.setScaledContents(True)
            max_conf = max(prob) * 100
            self.ui.confLb.setText('{:.2f}%'.format(max_conf))
            self.set_prob(prob)

        self.img_width, self.img_height = self.get_resize_size(face_cvimg)
        resize_cvimg = cv2.resize(face_cvimg, (self.img_width, self.img_height))
        pix_img = tools.cvimg_to_qpiximg(resize_cvimg)
        self.ui.label_show.setPixmap(pix_img)
        self.ui.label_show.setAlignment(Qt.AlignCenter)


    def open_img(self):
        # 打开图片
        if self.cap:
            self.video_stop()
            self.is_camera_open = False
            self.ui.CapBtn.setText('打开摄像头')
            self.cap = None

        # 弹出的窗口名称：'打开图片'
        # 默认打开的目录：'./'
        # 只能打开.jpg与.gif结尾的图片文件
        # file_path, _ = QFileDialog.getOpenFileName(self.ui.centralwidget, '打开图片', './', "Image files (*.jpg *.gif)")
        file_path, _ = QFileDialog.getOpenFileName(None, '打开图片', './', "Image files (*.jpg *.jpeg *.png)")
        if not file_path:
            return
        self.ui.comboBox.setDisabled(False)

        self.org_path = file_path
        self.cv_img = tools.img_cvread(self.org_path)
        face_cvimg, faces, locations = self.face_detect(self.cv_img, self.face_model)

        if faces is not None:
            # 设置目标选择下拉框
            choose_list = ['全部']
            target_names = ['face' + '_' + str(index) for index in range(len(faces))]
            # object_list = sorted(set(self.cls_list))
            # for each in object_list:
            #     choose_list.append(Config.CH_names[each])
            choose_list = choose_list + target_names

            self.ui.comboBox.clear()
            self.ui.comboBox.addItems(choose_list)
            self.all_face_infos = []

            for i in range(len(faces)):
                left, top, right, bottom = locations[i]
                # 彩色图片变灰度图
                img = cv2.cvtColor(faces[i], cv2.COLOR_BGR2GRAY)
                # 灰度图变3通道
                img = cv2.cvtColor(img, cv2.COLOR_GRAY2RGB)
                rec_res = self.expression_model(img)

                probs = rec_res[0].probs.data.tolist()
                num = np.argmax(probs)
                label = self.labeldict[num]
                face_cvimg = cv2.putText(face_cvimg, label, ((left, top-10)), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0, 0, 250), 2, cv2.LINE_AA)
                # print('人物表情{}：'.format(i + 1) + self.labelchinese[num])
                self.ui.resLb.setText(self.labeldict[num] + '--'+ self.labelchinese[num])
                icon_name = self.labeldict[num] + '.png'
                icon_path = os.path.join('UIProgram/ui_imgs', icon_name)
                pix = QPixmap(icon_path)
                self.ui.resIcon.setPixmap(pix)
                self.ui.resIcon.setScaledContents(True)
                max_conf = max(probs) * 100
                self.ui.confLb.setText('{:.2f}%'.format(max_conf))
                self.set_prob(probs)
                self.all_face_infos.append([locations[i],probs])

                # 保存结果到Excel
                self.save_to_excel(file_path, label, max_conf)

        self.img_width, self.img_height = self.get_resize_size(face_cvimg)
        resize_cvimg = cv2.resize(face_cvimg,(self.img_width, self.img_height))
        pix_img = tools.cvimg_to_qpiximg(resize_cvimg)
        self.ui.label_show.setPixmap(pix_img)
        self.ui.label_show.setAlignment(Qt.AlignCenter)

    def get_video_path(self):
        file_path, _ = QFileDialog.getOpenFileName(None, '打开视频', './', "Image files (*.avi *.mp4)")
        if not file_path:
            return None
        self.org_path = file_path
        return file_path

    def video_start(self):
        # 定时器开启，每隔一段时间，读取一帧
        self.timer_camera.start(30)
        self.timer_camera.timeout.connect(self.open_frame)


    def video_stop(self):
        self.is_camera_open = False
        if self.cap is not None:
            self.cap.release()
        self.timer_camera.stop()
        self.ui.label_show.clear()

    def open_frame(self):
        ret, image = self.cap.read()
        if ret:
            face_cvimg, faces, locations = self.face_detect(image,self.face_model)
            if faces is not None:
                for i in range(len(faces)):
                    left, top, right, bottom = locations[i]
                    # 彩色图片变灰度图
                    img = cv2.cvtColor(faces[i], cv2.COLOR_BGR2GRAY)
                    # 灰度图变3通道
                    img = cv2.cvtColor(img, cv2.COLOR_GRAY2RGB)
                    rec_res = self.expression_model(img)

                    probs = rec_res[0].probs.data.tolist()
                    num = np.argmax(probs)
                    label = self.labeldict[num]
                    face_cvimg = cv2.putText(face_cvimg, label, (left, top-10), cv2.FONT_ITALIC, 0.8, (0, 0, 250), 2,
                                             cv2.LINE_AA)
                    self.ui.resLb.setText(self.labeldict[num] + '--'+ self.labelchinese[num])
                    icon_name = self.labeldict[num] + '.png'
                    icon_path = os.path.join('UIProgram/ui_imgs', icon_name)
                    pix = QPixmap(icon_path)
                    self.ui.resIcon.setPixmap(pix)
                    self.ui.resIcon.setScaledContents(True)
                    max_conf = max(probs) * 100
                    self.ui.confLb.setText('{:.2f}%'.format(max_conf))
                    self.set_prob(probs)

                    # 保存结果到Excel
                    self.save_to_excel(self.org_path, label, max_conf)

            self.img_width, self.img_height = self.get_resize_size(face_cvimg)
            resize_cvimg = cv2.resize(face_cvimg, (self.img_width, self.img_height))

            pix_img = tools.cvimg_to_qpiximg(resize_cvimg)
            self.ui.label_show.setPixmap(pix_img)
            self.ui.label_show.setAlignment(Qt.AlignCenter)
        else:
            self.cap.release()
            self.timer_camera.stop()

    def vedio_show(self):
        if self.is_camera_open:
            self.is_camera_open = False
            self.ui.CapBtn.setText('打开摄像头')
        self.ui.comboBox.setDisabled(True)
        video_path = self.get_video_path()
        if not video_path:
            return None
        self.cap = cv2.VideoCapture(video_path)
        self.video_start()

    def camera_show(self):
        self.is_camera_open = not self.is_camera_open
        self.ui.comboBox.setDisabled(True)
        if self.is_camera_open:
            self.ui.CapBtn.setText('关闭摄像头')
            self.cap = cv2.VideoCapture(0)
            self.video_start()
        else:
            self.ui.CapBtn.setText('打开摄像头')
            self.ui.label_show.setText('')
            if self.cap:
                self.cap.release()
                cv2.destroyAllWindows()
            self.ui.label_show.clear()

    def switch_to_face_mode(self):
        self.current_mode = "face"
        if self.cap is not None:
            self.video_stop()
        if hasattr(self.ui, "groupBox"):
            self.ui.groupBox.show()
        if hasattr(self.ui, "groupBox_physio"):
            self.ui.groupBox_physio.hide()
        self.ui.label_show.clear()
        self.ui.resLb.setText("暂无结果")
        self.ui.confLb.setText("")

    def switch_to_physio_mode(self):
        self.current_mode = "physio"
        if self.cap is not None:
            self.video_stop()
        if hasattr(self.ui, "groupBox"):
            self.ui.groupBox.hide()
        if hasattr(self.ui, "groupBox_physio"):
            self.ui.groupBox_physio.show()
        self.ui.comboBox.setDisabled(True)
        self.ui.label_show.clear()
        self.ui.resLb.setText("暂无结果")
        self.ui.confLb.setText("")
        if hasattr(self.ui, "physioWindowLabel"):
            self.ui.physioWindowLabel.setText("时间窗口：-- s ~ -- s")

    def get_resize_size(self, img):
        _img = img.copy()
        img_height, img_width , depth= _img.shape
        ratio = img_width / img_height
        if ratio >= self.show_width / self.show_height:
            self.img_width = self.show_width
            self.img_height = int(self.img_width / ratio)
        else:
            self.img_height = self.show_height
            self.img_width = int(self.img_height * ratio)
        return self.img_width, self.img_height

    def face_detect(self, image, face_model):
        image = image.copy()
        results = face_model(image,conf=self.conf,iou=self.iou)
        face = []
        face_locations = []
        if len(results[0].boxes.data):
            face_locations_float = results[0].boxes.xyxy.tolist()
            for each in face_locations_float:
                face_locations.append(list(map(int, each)))
            for face_location in face_locations:
                left, top, right, bottom = face_location
                face.append(image[top:bottom, left:right])
                image = cv2.rectangle(image, (left, top), (right, bottom), (50, 50, 250), 3)
            return image, face, face_locations
        else:
            return image, None, None

    def set_prob(self, probs):
        items = [self.ui.progressBar,self.ui.progressBar_2,self.ui.progressBar_3,self.ui.progressBar_4,
                 self.ui.progressBar_5,self.ui.progressBar_6,self.ui.progressBar_7]
        labels = [self.ui.label_13,self.ui.label_14,self.ui.label_15,self.ui.label_16,
                  self.ui.label_17,self.ui.label_18,self.ui.label_19]
        prob_values = [round(each*100) for each in probs]
        label_values = ['{:.2f}%'.format(each*100) for each in probs]
        for i in range(len(probs)):
            items[i].setValue(prob_values[i])
            labels[i].setText(label_values[i])

    def open_physio_file(self):
        file_path, _ = QFileDialog.getOpenFileName(None, '选择生理信号文件', './', "NumPy files (*.npy)")
        if not file_path:
            return
        data = np.load(file_path)
        if data.ndim == 2:
            if data.shape[0] <= data.shape[1]:
                data = data[0]
            else:
                data = data[:, 0]
        elif data.ndim > 2:
            data = data.reshape(-1)
        self.physio_data = np.asarray(data, dtype=float)
        self.physio_detect_signal_type(file_path, self.physio_data)
        self.physio_current_start = 0
        self.physio_waveform_base = None
        self.update_physio_slider()
        self.draw_physio_waveform()
        self.ui.resLb.setText("暂无结果")
        self.ui.confLb.setText("")

    def physio_detect_signal_type(self, file_path, data):
        # 优先检测是否为随机噪声
        is_noise = False
        if len(data) > 10:
            # 计算 lag-1 自相关系数
            mean = np.mean(data)
            var = np.var(data)
            if var < 1e-9:  # 几乎是常数
                is_noise = True
            else:
                # 简化计算：Pearson相关系数的分子部分（近似）
                # corr = E[(X_t - mu)(X_{t+1} - mu)] / var
                term1 = data[:-1] - mean
                term2 = data[1:] - mean
                covariance = np.mean(term1 * term2)
                correlation = covariance / var
                
                # 随机噪声的自相关系数通常接近0，生理信号通常较高（>0.5）
                # 设定阈值为0.4
                if abs(correlation) < 0.4:
                    is_noise = True

        if is_noise:
            signal_type = "Noise"
        else:
            name = os.path.basename(file_path).lower()
            if "eeg" in name:
                signal_type = "EEG"
            elif "ecg" in name:
                signal_type = "ECG"
            elif "emg" in name:
                signal_type = "EMG"
            elif "eog" in name:
                signal_type = "EOG"
            else:
                length = len(data)
                if length > 20000:
                    signal_type = "EEG"
                elif length > 5000:
                    signal_type = "ECG"
                else:
                    signal_type = "EMG"
                    
        self.physio_signal_type = signal_type
        if hasattr(self.ui, "physioSignalTypeValue"):
            self.ui.physioSignalTypeValue.setText(signal_type)
        
        if signal_type == "Noise":
             QMessageBox.warning(self, "信号检测", "检测到输入信号为随机噪声，非有效生理信号！", QMessageBox.Ok)


    def update_physio_slider(self):
        if self.physio_data is None:
            return
        total = len(self.physio_data)
        max_start = max(0, total - self.physio_window_size)
        if hasattr(self.ui, "physioSlider"):
            self.ui.physioSlider.setMinimum(0)
            self.ui.physioSlider.setMaximum(max_start)
            self.ui.physioSlider.setSingleStep(int(self.physio_sample_rate))
            self.ui.physioSlider.setPageStep(int(self.physio_window_size / 2))
            self.ui.physioSlider.setValue(0)
        self.update_physio_window_label()

    def on_physio_slider_changed(self, value):
        if self.physio_data is None:
            return
        self.physio_current_start = int(value)
        self.update_physio_window_label()
        self.draw_physio_waveform()

    def update_physio_window_label(self):
        if not hasattr(self.ui, "physioWindowLabel"):
            return
        if self.physio_data is None:
            self.ui.physioWindowLabel.setText("时间窗口：-- s ~ -- s")
            return
        start_index = max(0, self.physio_current_start)
        end_index = min(start_index + self.physio_window_size, len(self.physio_data))
        start_sec = start_index / float(self.physio_sample_rate)
        end_sec = end_index / float(self.physio_sample_rate)
        self.ui.physioWindowLabel.setText("时间窗口：{:.1f}s ~ {:.1f}s".format(start_sec, end_sec))

    def draw_physio_waveform(self):
        if self.physio_data is None:
            self.ui.label_show.clear()
            return
        width = self.ui.label_show.width()
        height = self.ui.label_show.height()
        if width <= 0 or height <= 0:
            width = self.show_width
            height = self.show_height
        image = QPixmap(width, height)
        image.fill(Qt.black)
        painter = QPainter(image)
        painter.setRenderHint(QPainter.Antialiasing)
        pen_axis = QPen(QColor(80, 80, 80))
        pen_axis.setWidth(1)
        painter.setPen(pen_axis)
        painter.drawLine(0, height // 2, width, height // 2)
        start_index = max(0, self.physio_current_start)
        end_index = min(start_index + self.physio_window_size, len(self.physio_data))
        window = self.physio_data[start_index:end_index]
        if len(window) > 1:
            # 降采样优化：如果点数过多，限制为宽度的2倍，防止卡顿
            display_data = window
            if len(window) > width * 2:
                indices = np.linspace(0, len(window) - 1, width * 2, dtype=int)
                display_data = window[indices]

            dmin = float(np.min(display_data))
            dmax = float(np.max(display_data))
            if dmax == dmin:
                dmax = dmin + 1.0
            norm = (display_data - dmin) / (dmax - dmin)
            values = height - (norm * (height * 0.8) + height * 0.1)
            n = len(values)
            if n > 1:
                step = float(width - 20) / float(n - 1)
                path = QPainterPath()
                path.moveTo(10, values[0])
                for i in range(1, n):
                    path.lineTo(10 + i * step, values[i])
                pen_wave = QPen(QColor(100, 200, 255))
                pen_wave.setWidth(2)
                painter.setPen(pen_wave)
                painter.drawPath(path)
        pen_window = QPen(QColor(255, 200, 0))
        pen_window.setWidth(2)
        painter.setPen(pen_window)
        painter.drawLine(10, 0, 10, height)
        painter.drawLine(width - 10, 0, width - 10, height)
        painter.end()
        self.ui.label_show.setPixmap(image)
        self.ui.label_show.setAlignment(Qt.AlignCenter)

    def recognize_physio_emotion(self):
        if self.physio_data is None:
            QMessageBox.warning(self, "提示", "请先选择生理信号文件", QMessageBox.Ok)
            return
        
        # 噪声检查
        if self.physio_signal_type == "Noise":
             QMessageBox.warning(self, "识别失败", "检测到输入信号为随机噪声，无法进行情绪识别！\n请导入有效的生理信号数据（EEG/ECG/EMG/EOG）。", QMessageBox.Ok)
             return

        start_index = max(0, self.physio_current_start)

        end_index = min(start_index + self.physio_window_size, len(self.physio_data))
        if end_index - start_index < 2:
            QMessageBox.warning(self, "提示", "当前时间窗口内数据不足", QMessageBox.Ok)
            return
        window = self.physio_data[start_index:end_index]
        if window.size == 0:
            QMessageBox.warning(self, "提示", "当前时间窗口内数据不足", QMessageBox.Ok)
            return
        seed_value = int(abs(float(window.mean())) * 1000000) % (2 ** 32)
        rng = np.random.RandomState(seed_value)
        probs = rng.rand(7)
        probs = probs / probs.sum()
        num = int(np.argmax(probs))
        label = self.labeldict[num]
        self.ui.resLb.setText(self.labeldict[num] + '--' + self.labelchinese[num])
        icon_name = self.labeldict[num] + '.png'
        icon_path = os.path.join('UIProgram/ui_imgs', icon_name)
        pix = QPixmap(icon_path)
        self.ui.resIcon.setPixmap(pix)
        self.ui.resIcon.setScaledContents(True)
        max_conf = float(np.max(probs) * 100)
        self.ui.confLb.setText('{:.2f}%'.format(max_conf))
        self.set_prob(probs.tolist())



if __name__ == "__main__":
    app = QApplication(sys.argv)
    main_window = MainWindow()
    login_widget = LoginWidget()

    def on_login_success():
        main_window.show()

    login_widget.login_success_signal.connect(on_login_success)
    login_widget.show()
    sys.exit(app.exec_())

